import xlrd,math
from prettytable import PrettyTable
from openpyxl import load_workbook



#	      序号   药品名        品牌  单价（元）    	 规格	          每盒粒数	      每天服用次数	每次服用片数  服用时间	饭前/饭后	剩余粒数

title = ['num','drug_name','brand','price','Specifications','each_box_grain','evDay_eat_num','each_time_piece','eat_time','before_after','inventory']
medi_db = []
eat_time_List = ['早','中','晚']
before_after_List = ['饭前','饭后']

def read_excel():
    xlsx = xlrd.open_workbook('db.xlsx',formatting_info = False)
    sheet = xlsx.sheet_by_name("Sheet1")
    print('表名称:%s,表行数:%s,表列数:%s'% (sheet.name,sheet.nrows,sheet.ncols))

    for nrow in range(1,sheet.nrows):
        temp_dic = {}
        for ncol,title_n in zip(range(sheet.ncols),title):
            c_value = sheet.cell_value(nrow,ncol)
            temp_dic.update({title_n:c_value})
        medi_db.append(temp_dic)
    #print(medi_db)

def buy_calc(days):
    wb = load_workbook('db.xlsx')
    ws =wb['buy']
    wb.remove(ws)
    ws = wb.create_sheet("buy")
    table_title = ["序号", "药品名称","品牌","单价","规格","每盒粒数", "该药品总价", "需采购盒数"]
    x= PrettyTable(table_title)
    ws.append(table_title)
    count_price = 0
    for drug,index in zip(medi_db,range(len(medi_db))):

        each_box_grain = int(drug['each_box_grain'])
        evDay_eat_num =  int(drug['evDay_eat_num'])
        each_time_piece = int(drug['each_time_piece'])
        inventory = int(drug['inventory'])
        box_num = math.ceil( (int(days) * evDay_eat_num * each_time_piece - inventory)/each_box_grain )   #备药天数/(每盒粒数/(每天吃的粒数*每顿吃的片数))
        per_case_price = box_num * drug['price']
        result_list = [index+1,drug['drug_name'],drug['brand'],drug['price'],drug['Specifications'],each_box_grain,per_case_price,box_num]
        x.add_row(result_list)
        ws.append(result_list)
        count_price += per_case_price
    print(f'\r\n\r\n备药天数:{days}天，购药总价{count_price}元')
    print(x)
    wb.save('db.xlsx')

eat_time_List = ['早','中','晚']
before_after_List = ['饭前','饭后']

def eat_sch():
    wb = load_workbook('db.xlsx')
    ws =wb['take_medi_sch']
    wb.remove(ws)
    ws = wb.create_sheet("take_medi_sch")
    table_title = ["早/中/晚", "饭前/饭后", "药品名称", "服药片数（片/粒）"]
    x= PrettyTable(table_title)
    ws.append(table_title)
    for timev,index in zip(eat_time_List,range(len(eat_time_List))):
        #if timev in eat_time_list:
            for b_a in before_after_List:
                for drug,ind in zip(medi_db,range(len(medi_db))):
                    eat_time_list = drug['eat_time'].split(',')
                    before_after = drug['before_after']
                    if timev in eat_time_list and before_after == b_a:
                        #time_n = f"{timev}_{b_a}"

                        each_time_piece = int(drug['each_time_piece'])
                        result_list = [timev,b_a,drug['drug_name'],each_time_piece]
                        x.add_row(result_list)
                        ws.append(result_list)
            x.add_row(['','','',''])
    x.del_row(-1)
    print('\r\n\r\n服药安排：')
    print(x)
    wb.save('db.xlsx')




def input_top():
    print('''
【1】根据输入的备药天数，计算各种药品需要采购的数量
【2】输出服药时间计划
    ''')
    fun_num= input('请输入数字，并回车：')
    input_down(fun_num)

def input_down(num):
    if num == '1':
        days = input('请输入需要备药的天数：')
        buy_calc(days)
    elif num == '2':
        eat_sch()

if __name__ == '__main__':
    read_excel()
    input_top()
        #hostid=sheet.row_values(i)[0]
